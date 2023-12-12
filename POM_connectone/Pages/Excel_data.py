import openpyxl

def getrowcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row

def getcolumncount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column

def getcelldata(path, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row =rownum,column=colnum).value

def setcelldata(path, sheetname, rownum, colnum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row =rownum,column=colnum).value=data
    workbook.save(path)

path = "C:\\Users\Shantanu\PycharmProjects\Connectone\Excel\Test Data.xlsx"
sheetname = "Sheet1"
rows = getrowcount(path, sheetname)
columns = getcolumncount(path, sheetname)

print(rows, "---", columns)
print(getcelldata(path, sheetname, 1, 7))