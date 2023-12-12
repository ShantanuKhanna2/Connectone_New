import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

import pytest
Chrome_options = Options()


@pytest.fixture(params=["chrome"],scope="class")
def get_browser(request):
    global web_driver
    if request.param == "chrome":
        web_driver = webdriver.Chrome(executable_path="C:\\Users\Shantanu\PycharmProjects\Connectone\Driver\chromedriver.exe")
    request.cls.driver = web_driver
    web_driver.maximize_window()
    web_driver.implicitly_wait(10)
    yield
    web_driver.quit()

def get_data():
    workbook = openpyxl.load_workbook("C:\\Users\Shantanu\PycharmProjects\Connectone\Excel\Test Data.xlsx")
    sheet = workbook["Sheet1"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
           data = sheet.cell(row=i,column=j).value
           dataList.insert(j,data)
        mainList.insert(i,dataList)
    return mainList


@pytest.mark.usefixtures("get_browser")
class BasesTest:
    pass


class TestHubSpot(BasesTest):
    @pytest.mark.parametrize(
        "first_name,start_date,end_date,suffix,start_msg,end_msg,question_data,answer_mode", get_data())
    def test_dologin(self, first_name,start_date,end_date,suffix,start_msg,end_msg,question_data,answer_mode):
        driver = self.driver
        driver.get("https://connectonetest2.bkinfo.in/Start/Login/Frm_Auditor_Login")
        wait = WebDriverWait(driver, 20)
        username = wait.until(EC.element_to_be_clickable((By.ID, "Txt_User")))
        ActionChains(driver).move_to_element(username).send_keys("bksatendra").perform()

        password = wait.until(EC.element_to_be_clickable((By.NAME, "Txt_Pass")))
        password.send_keys("babatest")

        login_click = wait.until(EC.element_to_be_clickable((By.ID, "Login")))
        ActionChains(driver).click(login_click).perform()

        # swtich to new window
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element(By.XPATH, "//*[@id='CenterSelectionAuditorListGrid_DXDataRow0']/td[1]").click()
        driver.find_element(By.XPATH, "//*[@id='SelectInstitute_Save']/div").click()
        driver.find_element(By.XPATH, "//*[@id='COD_Selection_Auditor_Grid_DXDataRow0']/td[2]").click()
        driver.find_element(By.XPATH, "//*[@id='SelectCenter_Save']/div").click()

        # For Stale element exception error
        ActionChains(driver).double_click(driver.find_element(By.XPATH, "//*[@id='SelectCenter_Save']/div")).perform()
        time.sleep(5)

        # For dropdown, action chain is used if index, visible_text,etc are not working.
        driver.find_element(By.ID, "Facility_Main").click()
        drop_down = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='Facility_GodlyServices']")))
        ActionChains(driver).click(drop_down).perform()
        drop_down_1 = wait.until(EC.element_to_be_clickable((By.ID, "Facility_ChartInfo")))
        ActionChains(driver).click(drop_down_1).perform()
        time.sleep(5)

        # create form
        driver.find_element(By.XPATH, "//*[@id='ChartInfoCreateForm']/div/span").click()
        time.sleep(10)

        # data-driver starts from here:
        form_name = wait.until(EC.presence_of_element_located((By.ID, "FormName_createForm")))
        ActionChains(driver).move_to_element(form_name).click().send_keys("Form no.1").perform()


        start_date = driver.find_element(By.ID, "StartDate_createForm")
        ActionChains(driver).move_to_element(start_date).click().send_keys("28-01-2023").perform()

        end_date = driver.find_element(By.ID, "EndDate_createForm")
        ActionChains(driver).move_to_element(end_date).click().send_keys("31-01-2023").perform()

        purpose = wait.until(EC.element_to_be_clickable((By.ID, "Purpose_createForm")))
        ActionChains(driver).move_to_element(purpose).click().send_keys("BASIC DETAILS").perform()
        purpose.send_keys(Keys.ENTER)
        time.sleep(5)

        suffix = wait.until(EC.presence_of_element_located((By.ID, "RegNoPrefix_createForm")))
        ActionChains(driver).move_to_element(suffix).click().send_keys("H").perform()

        start_msg = driver.find_element(By.XPATH, "//*[@id='StartDateMsg_createForm']/div[2]/div[1]/p")
        ActionChains(driver).move_to_element(start_msg).click().send_keys("Hello everyone").perform()

        end_msg = driver.find_element(By.XPATH, "//*[@id='EndDateMsg_createForm']/div[2]/div[1]/p")
        ActionChains(driver).move_to_element(end_msg).click().send_keys("Thank you").perform()

        add_question = driver.find_element(By.XPATH, "//*[@id='devextreme134']/div")
        ActionChains(driver).move_to_element(add_question).click(add_question).perform()

        question_1 = wait.until(EC.element_to_be_clickable((By.ID, "Section_0__Question")))
        ActionChains(driver).move_to_element(question_1).click().send_keys("Tell me about yourself").perform()

        answer_mode = wait.until(EC.element_to_be_clickable((By.ID, "Section_0__Mode")))
        # answer_mode = driver.find_element(By.XPATH,"//*[@id='dx-2a582bd4-5578-6155-bcd3-c44f75926cb4']/div[1]/div/div[1]/div[2]/div[1]/div")
        ActionChains(driver).move_to_element(answer_mode).click().send_keys("Short Answer").perform()
        answer_mode.send_keys(Keys.ENTER)
        time.sleep(5)

        save_form = driver.find_element(By.XPATH, "//*[@id='SaveBtn_createForm']/div")
        ActionChains(driver).move_to_element(save_form).click(save_form).perform()
        time.sleep(15)

