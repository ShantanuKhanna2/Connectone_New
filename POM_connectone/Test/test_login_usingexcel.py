import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

import pytest
Chrome_options = Options()


@pytest.fixture(params=["chrome"],scope="class")
def get_browser(request):
    global web_driver
    if request.param == "chrome":
        web_driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
    request.cls.driver = web_driver
    web_driver.maximize_window()
    web_driver.implicitly_wait(10)
    yield
    web_driver.quit()

def get_data():
    workbook = openpyxl.load_workbook("C:\\Users\Shantanu\PycharmProjects\Connectone\Excel\Book1.xlsx")
    sheet = workbook["Sheet1"]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []

    for i in range(2,totalrows+1):
        dataList = []
        for j in range(1, totalcols+1):
           data = sheet.cell(row=i,column=j).value
           dataList.insert(j,data)
        mainList.insert(i,dataList)
    return mainList


@pytest.mark.usefixtures("get_browser")
class BasesTest:
    pass


class TestHubSpot(BasesTest):
    @pytest.mark.parametrize(
        "username, password", get_data())
    def test_dologin(self, username, password):
        driver = self.driver
        driver.get("https://connectonetest2.bkinfo.in/Start/Login/Frm_Auditor_Login")
        wait = WebDriverWait(driver, 20)
        username_1 = wait.until(EC.element_to_be_clickable((By.ID, "Txt_User")))
        ActionChains(driver).move_to_element(username_1).send_keys(username).perform()

        password_1 = wait.until(EC.element_to_be_clickable((By.NAME, "Txt_Pass")))
        password_1.send_keys(password)

        login_click = wait.until(EC.element_to_be_clickable((By.ID, "Login")))
        ActionChains(driver).click(login_click).perform()
